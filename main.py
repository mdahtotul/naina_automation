import openpyxl

inv_file = openpyxl.load_workbook("inventory.xlsx")
product_list = inv_file["Sheet1"]

products_per_supplier = {}
total_value_per_supplier = {}
low_product_inv = {}

total_row = product_list.max_row + 1
product_list.cell(1, 5).value = "Inventory Price"

# list each company with respective product count
for product_row in range(2, total_row):
    supplier_name = product_list.cell(product_row, 4).value
    inventory = product_list.cell(product_row, 2).value
    price = product_list.cell(product_row, 3).value
    product_num = product_list.cell(product_row, 1).value
    inventory_price = product_list.cell(product_row, 5)

    # calculation of number of products per supplier
    if supplier_name in products_per_supplier:
        current_num_products = products_per_supplier.get(supplier_name)
        products_per_supplier[supplier_name] = current_num_products + 1
    else:
        products_per_supplier[supplier_name] = 1

    # calculation of total value of inventory per supplier
    if supplier_name in total_value_per_supplier:
        current_total_value = total_value_per_supplier.get(supplier_name)
        total_value_per_supplier[supplier_name] = current_total_value + (inventory * price)
    else:
        total_value_per_supplier[supplier_name] = inventory * price

    # logic for inventory product less than 10
    if inventory < 10:
        low_product_inv[product_num] = inventory

    # add value for total inventory price
    inventory_price.value = inventory * price

print(products_per_supplier)
print(total_value_per_supplier)
print(low_product_inv)

inv_file.save("inventory_with_total_value.xlsx")
